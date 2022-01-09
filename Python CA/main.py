'''
this program edits and update emails in an excel file from
@helpinghands.cm to  @handsinhands.org
'''

import openpyxl as xl

wb = xl.load_workbook('employeedata.xlsx')
sheet = wb['Sheet1'] # sheet 1 is the active page that is why it is called in the workbook(wb)
#cell = sheet['a1']
cell = sheet.cell(2,  2)# reads the data in row 2 column 2
#cell = sheet.cell(row = 2, column = 2)#

print(cell.value)

'''
# ------- defining the rows and column ----#
row = sheet.max_row
column = sheet.max_column   '''

# ----- creating a new row for updated email -----#
for row in range (2, sheet.max_row + 1):# begins from two because we're working with the second row elements

    cell = sheet.cell(row, 1 )# reads the element of column 1 in all the rows

    new_email = (cell.value + '@helpinghands.cm') # reads the element of the cell e.g jude and add @helpinghands.cm to it
    #----- creating a new cell to store the new email -----#

    new_email_cell = sheet.cell(row, 4) # stores the new email in the 4th column(column D)

    #------ attributing the data to the new cells (column) created -----#
    new_email_cell.value = new_email
    
# ----- creating a new database with the .csv extention that stores the new and old email 
wb.save('updated_emails.csv')

old_domain = 'helpinghands.cm'
new_domain = 'handsinhands.org'
for i in range(2,sheet.max_row + 1): 
    cell = sheet.cell(i,2) 
    if old_domain in cell.value:
        updated_email= cell.value.replace(old_domain, new_domain) 

        sheet.cell(i,2).value = updated_email
wb.save('updated_emails.csv')





